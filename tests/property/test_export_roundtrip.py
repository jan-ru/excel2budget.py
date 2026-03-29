"""Property tests for export round-trip.

**Property 10: Export round-trip** — export and re-parse preserves
column ordering, row count, and data values.

**Validates: Requirements 9.1, 9.2, 9.3, 9.4**
"""

from __future__ import annotations

import csv
import io

from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook

from src.core.types import (
    ColumnDef,
    DataMetadata,
    DataType,
    FloatVal,
    IntVal,
    NullVal,
    OutputTemplate,
    Row,
    StringVal,
    TabularData,
    TemplateColumnDef,
    FromSource,
)
from src.export.exporter import exportToCSV, exportToExcel


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

_col_name = st.from_regex(r"[A-Za-z][A-Za-z0-9_]{0,14}", fullmatch=True)


@st.composite
def tabular_data_with_template(draw: st.DrawFn):
    """Generate TabularData with STRING, INTEGER, and FLOAT columns.

    Also returns a matching OutputTemplate for Excel export.
    NullVal cells are excluded to keep round-trip comparison simple.
    """
    num_cols = draw(st.integers(min_value=1, max_value=8))
    num_rows = draw(st.integers(min_value=0, max_value=15))

    col_names = draw(
        st.lists(_col_name, min_size=num_cols, max_size=num_cols, unique=True)
    )
    col_types = draw(
        st.lists(
            st.sampled_from([DataType.STRING, DataType.INTEGER, DataType.FLOAT]),
            min_size=num_cols,
            max_size=num_cols,
        )
    )

    columns = [
        ColumnDef(name=n, dataType=t) for n, t in zip(col_names, col_types)
    ]

    rows = []
    for _ in range(num_rows):
        values = []
        for dt in col_types:
            if dt == DataType.STRING:
                # Avoid newlines/commas/quotes that complicate CSV round-trip
                v = draw(st.from_regex(r"[A-Za-z0-9 ]{0,20}", fullmatch=True))
                values.append(StringVal(v))
            elif dt == DataType.INTEGER:
                v = draw(st.integers(min_value=-999999, max_value=999999))
                values.append(IntVal(v))
            else:  # FLOAT
                v = draw(
                    st.floats(
                        min_value=-99999,
                        max_value=99999,
                        allow_nan=False,
                        allow_infinity=False,
                    ).map(lambda x: round(x, 4))
                )
                values.append(FloatVal(v))
        rows.append(Row(values=values))

    data = TabularData(
        columns=columns,
        rows=rows,
        rowCount=num_rows,
        metadata=DataMetadata(),
    )

    template = OutputTemplate(
        packageName="test",
        templateName="roundtrip",
        columns=[
            TemplateColumnDef(
                name=c.name,
                dataType=c.dataType,
                nullable=True,
                sourceMapping=FromSource(sourceColumnName=c.name),
            )
            for c in columns
        ],
    )

    return data, template, col_types


# ---------------------------------------------------------------------------
# CSV round-trip
# ---------------------------------------------------------------------------

@given(data=tabular_data_with_template())
@settings(max_examples=50, deadline=None)
def test_property_10_csv_roundtrip_preserves_columns_and_rows(data):
    """CSV export and re-parse preserves column ordering and row count."""
    tabular, _template, col_types = data

    csv_bytes = exportToCSV(tabular)
    assert isinstance(csv_bytes, bytes)

    # Re-parse
    reader = csv.reader(io.StringIO(csv_bytes.decode("utf-8")))
    parsed_rows = list(reader)

    # Header preserved
    assert len(parsed_rows) >= 1, "CSV must have at least a header row"
    header = parsed_rows[0]
    expected_names = [c.name for c in tabular.columns]
    assert header == expected_names, "Column ordering must be preserved"

    # Row count preserved
    data_rows = parsed_rows[1:]
    assert len(data_rows) == tabular.rowCount, "Row count must be preserved"

    # Values preserved (compare as strings since CSV is text)
    for row_idx, (orig_row, csv_row) in enumerate(
        zip(tabular.rows, data_rows)
    ):
        assert len(csv_row) == len(tabular.columns), (
            f"Row {row_idx} column count mismatch"
        )
        for col_idx, (orig_cell, csv_val) in enumerate(
            zip(orig_row.values, csv_row)
        ):
            expected = _cell_to_csv_str(orig_cell)
            assert csv_val == expected, (
                f"Row {row_idx}, Col {col_idx}: "
                f"expected {expected!r}, got {csv_val!r}"
            )

    # exportedAt must be stamped
    assert tabular.metadata.exportedAt is not None


def _cell_to_csv_str(cell) -> str:
    """Convert a CellValue to the string representation CSV would produce."""
    if isinstance(cell, NullVal):
        return ""
    if isinstance(cell, StringVal):
        return cell.value
    if isinstance(cell, IntVal):
        return str(cell.value)
    if isinstance(cell, FloatVal):
        return str(cell.value)
    return ""


# ---------------------------------------------------------------------------
# Excel round-trip
# ---------------------------------------------------------------------------

@given(data=tabular_data_with_template())
@settings(max_examples=50, deadline=None)
def test_property_10_excel_roundtrip_preserves_columns_and_rows(data):
    """Excel export and re-parse preserves column ordering, row count, values."""
    tabular, template, col_types = data

    xlsx_bytes = exportToExcel(tabular, template)
    assert isinstance(xlsx_bytes, bytes)

    # Re-parse with openpyxl
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Header preserved
    header = [cell.value for cell in ws[1]]
    expected_names = [c.name for c in tabular.columns]
    assert header == expected_names, "Column ordering must be preserved"

    # Row count preserved
    actual_data_rows = ws.max_row - 1  # subtract header
    assert actual_data_rows == tabular.rowCount, "Row count must be preserved"

    # Values preserved
    for row_idx, orig_row in enumerate(tabular.rows):
        xlsx_row = [cell.value for cell in ws[row_idx + 2]]
        for col_idx, (orig_cell, xlsx_val) in enumerate(
            zip(orig_row.values, xlsx_row)
        ):
            expected = _cell_to_python(orig_cell)
            if isinstance(orig_cell, FloatVal):
                # Float comparison with tolerance
                assert xlsx_val is not None, (
                    f"Row {row_idx}, Col {col_idx}: expected float, got None"
                )
                assert abs(float(xlsx_val) - expected) < 1e-9, (
                    f"Row {row_idx}, Col {col_idx}: "
                    f"expected {expected}, got {xlsx_val}"
                )
            elif isinstance(orig_cell, StringVal) and orig_cell.value == "":
                # openpyxl reads empty strings back as None
                assert xlsx_val is None or xlsx_val == "", (
                    f"Row {row_idx}, Col {col_idx}: "
                    f"expected empty/None, got {xlsx_val!r}"
                )
            else:
                assert xlsx_val == expected, (
                    f"Row {row_idx}, Col {col_idx}: "
                    f"expected {expected!r}, got {xlsx_val!r}"
                )

    # exportedAt must be stamped
    assert tabular.metadata.exportedAt is not None

    wb.close()


def _cell_to_python(cell):
    """Convert CellValue to plain Python value."""
    if isinstance(cell, NullVal):
        return None
    if isinstance(cell, StringVal):
        return cell.value
    if isinstance(cell, IntVal):
        return cell.value
    if isinstance(cell, FloatVal):
        return cell.value
    return None
