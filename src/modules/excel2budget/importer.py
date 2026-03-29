"""Excel budget file importer.

Parses .xlsx files, extracts budget data and column mapping configuration.
Detects month columns using Dutch month name conventions (jan, feb, mrt, ...).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.core.types import (
    CellValue,
    ColumnDef,
    DataMetadata,
    DataType,
    FileFormat,
    FloatVal,
    IntVal,
    MappingConfig,
    MonthColumnDef,
    NullVal,
    Row,
    StringVal,
    TabularData,
)


# --- Error types ---


@dataclass
class ParseError:
    """Error when the file cannot be parsed as a valid .xlsx or lacks expected sheets."""
    message: str
    available_sheets: List[str] = field(default_factory=list)


@dataclass
class MappingError:
    """Error when required mapping columns cannot be identified."""
    message: str
    missing_columns: List[str] = field(default_factory=list)
    available_columns: List[str] = field(default_factory=list)


# --- Constants ---

# Dutch month abbreviations used in budget Excel files (index 0 = January)
DUTCH_MONTHS: List[str] = [
    "jan", "feb", "mrt", "apr", "mei", "jun",
    "jul", "aug", "sep", "okt", "nov", "dec",
]

# Required fixed columns in the budget data
REQUIRED_COLUMNS = {"Entity", "Account", "DC"}

# Pattern to match month column names like "jan-26", "feb-24", "mrt-2026"
MONTH_COLUMN_PATTERN = re.compile(
    r"^(" + "|".join(DUTCH_MONTHS) + r")-(\d{2,4})$", re.IGNORECASE
)


# --- Public API ---


def parseExcelFile(raw_bytes: bytes) -> Workbook | ParseError:
    """Parse raw bytes into an openpyxl Workbook.

    Returns a ParseError if the bytes are not a valid .xlsx file.
    """
    try:
        wb = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
        return wb
    except Exception as exc:
        return ParseError(
            message=f"Expected a valid .xlsx file, but parsing failed: {exc}",
            available_sheets=[],
        )


def extractBudgetData(
    workbook: Workbook, sheet_name: str = "Budget"
) -> TabularData | ParseError:
    """Extract budget data from the named sheet.

    Returns a ParseError if the sheet does not exist, listing available sheets.
    """
    available = workbook.sheetnames
    if sheet_name not in available:
        return ParseError(
            message=f"Sheet '{sheet_name}' not found in workbook",
            available_sheets=available,
        )

    ws = workbook[sheet_name]
    rows_iter = ws.iter_rows()

    # First row is the header
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return ParseError(
            message=f"Sheet '{sheet_name}' is empty",
            available_sheets=available,
        )

    col_names = [str(cell.value).strip() if cell.value is not None else f"_col{i}" for i, cell in enumerate(header_row)]
    columns = [ColumnDef(name=name, dataType=DataType.STRING) for name in col_names]

    data_rows: List[Row] = []
    for row in rows_iter:
        values: List[CellValue] = []
        for cell in row:
            values.append(_cell_to_value(cell.value))
        # Pad or trim to match column count
        while len(values) < len(columns):
            values.append(NullVal())
        values = values[: len(columns)]
        data_rows.append(Row(values=values))

    return TabularData(
        columns=columns,
        rows=data_rows,
        rowCount=len(data_rows),
        metadata=DataMetadata(sourceName=sheet_name, sourceFormat=FileFormat.EXCEL),
    )


def extractMappingConfig(
    workbook: Workbook, sheet_name: str = "Budget"
) -> MappingConfig | MappingError:
    """Extract column mapping configuration from the budget sheet.

    Identifies Entity, Account, and DC columns, plus month columns.
    Returns a MappingError if required columns cannot be found.
    """
    available = workbook.sheetnames
    if sheet_name not in available:
        return MappingError(
            message=f"Sheet '{sheet_name}' not found",
            available_columns=[],
        )

    ws = workbook[sheet_name]
    try:
        header_row = next(ws.iter_rows(max_row=1))
    except StopIteration:
        return MappingError(message=f"Sheet '{sheet_name}' is empty")

    col_names = [
        str(cell.value).strip() if cell.value is not None else f"_col{i}"
        for i, cell in enumerate(header_row)
    ]

    # Find required columns (case-sensitive match)
    missing: List[str] = []
    found: Dict[str, str] = {}
    for req in REQUIRED_COLUMNS:
        match = _find_column(req, col_names)
        if match is None:
            missing.append(req)
        else:
            found[req] = match

    if missing:
        return MappingError(
            message=f"Required columns not found: {', '.join(sorted(missing))}",
            missing_columns=sorted(missing),
            available_columns=col_names,
        )

    # Detect month columns
    month_cols = _detect_month_columns_from_headers(col_names)
    if not month_cols:
        return MappingError(
            message="No month columns detected. Expected columns like 'jan-26', 'feb-26', etc.",
            available_columns=col_names,
        )

    return MappingConfig(
        entityColumn=found["Entity"],
        accountColumn=found["Account"],
        dcColumn=found["DC"],
        monthColumns=month_cols,
    )


def detectMonthColumns(
    data: TabularData, config: MappingConfig
) -> List[MonthColumnDef] | MappingError:
    """Validate that month columns referenced in config exist in the data.

    Returns a MappingError if any referenced month column is missing.
    """
    col_names = {col.name for col in data.columns}
    missing = [
        mc.sourceColumnName
        for mc in config.monthColumns
        if mc.sourceColumnName not in col_names
    ]
    if missing:
        return MappingError(
            message=f"Month columns not found in data: {', '.join(missing)}",
            missing_columns=missing,
            available_columns=[col.name for col in data.columns],
        )
    return config.monthColumns


# --- Internal helpers ---


def _cell_to_value(raw) -> CellValue:
    """Convert a raw cell value to a typed CellValue."""
    if raw is None:
        return NullVal()
    if isinstance(raw, bool):
        # Must check bool before int since bool is a subclass of int
        return StringVal(value=str(raw))
    if isinstance(raw, int):
        return IntVal(value=raw)
    if isinstance(raw, float):
        return FloatVal(value=raw)
    return StringVal(value=str(raw))


def _find_column(required: str, col_names: List[str]) -> Optional[str]:
    """Find a required column by exact match first, then case-insensitive."""
    for name in col_names:
        if name == required:
            return name
    for name in col_names:
        if name.lower() == required.lower():
            return name
    return None


def _detect_month_columns_from_headers(
    col_names: List[str],
) -> List[MonthColumnDef]:
    """Detect month columns from header names using the Dutch month pattern.

    Matches patterns like 'jan-26', 'feb-2026', 'mrt-24'.
    Returns a list of MonthColumnDef sorted by period number.
    """
    results: List[MonthColumnDef] = []
    for name in col_names:
        match = MONTH_COLUMN_PATTERN.match(name.strip())
        if match:
            month_abbr = match.group(1).lower()
            year_str = match.group(2)
            period = DUTCH_MONTHS.index(month_abbr) + 1  # 1-based
            year = _normalize_year(int(year_str))
            results.append(
                MonthColumnDef(
                    sourceColumnName=name.strip(),
                    periodNumber=period,
                    year=year,
                )
            )
    # Sort by period number for consistent ordering
    results.sort(key=lambda mc: mc.periodNumber)
    return results


def _normalize_year(year: int) -> int:
    """Normalize 2-digit years to 4-digit (e.g., 26 -> 2026)."""
    if year < 100:
        return 2000 + year
    return year
