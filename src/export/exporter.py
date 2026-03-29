"""Format exporters for CSV and Excel output.

Serializes TabularData into CSV or Excel (.xlsx) format, preserving
column ordering from the OutputTemplate and stamping export metadata.

Requirements: 9.1, 9.2, 9.3, 9.4, 9.5, 19.3
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from openpyxl import Workbook

from src.core.types import (
    BoolVal,
    CellValue,
    DataType,
    DateVal,
    FloatVal,
    IntVal,
    NullVal,
    OutputTemplate,
    StringVal,
    TabularData,
)


def _cell_to_python(cell: CellValue):
    """Convert a CellValue to a plain Python value for serialization."""
    if isinstance(cell, NullVal):
        return None
    if isinstance(cell, StringVal):
        return cell.value
    if isinstance(cell, IntVal):
        return cell.value
    if isinstance(cell, FloatVal):
        return cell.value
    if isinstance(cell, BoolVal):
        return cell.value
    if isinstance(cell, DateVal):
        return cell.value
    return None


def exportToCSV(data: TabularData) -> bytes:
    """Serialize TabularData into a CSV file as bytes.

    Preserves column ordering from the TabularData columns.
    Stamps exportedAt in metadata.

    Returns UTF-8 encoded CSV bytes.
    """
    data.metadata.exportedAt = datetime.now(timezone.utc)

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header row
    writer.writerow([col.name for col in data.columns])

    # Data rows
    for row in data.rows:
        writer.writerow([_cell_to_python(v) for v in row.values])

    return buf.getvalue().encode("utf-8")


def exportToExcel(data: TabularData, template: OutputTemplate) -> bytes:
    """Serialize TabularData into an Excel (.xlsx) file as bytes.

    Preserves column ordering from the OutputTemplate.
    Stamps exportedAt in metadata.

    Returns .xlsx file bytes.
    """
    data.metadata.exportedAt = datetime.now(timezone.utc)

    wb = Workbook()
    ws = wb.active
    ws.title = template.templateName

    # Header row
    for col_idx, col_def in enumerate(data.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_def.name)

    # Data rows
    for row_idx, row in enumerate(data.rows, start=2):
        for col_idx, cell in enumerate(row.values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_to_python(cell))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
